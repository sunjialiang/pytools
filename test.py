"""Command-line tool for simple Excel operations using openpyxl."""

import argparse
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - dependency handling
    print("openpyxl is required. Install it with 'pip install openpyxl'.")
    sys.exit(1)

DEFAULT_FILE = "example.xlsx"


def create_workbook(path: str) -> None:
    """Create a new workbook with a header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Age"])
    wb.save(path)


def append_row(path: str, name: str, age: int) -> None:
    """Append a row to the workbook, creating it if needed."""
    if not os.path.exists(path):
        create_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([name, age])
    wb.save(path)


def print_rows(path: str) -> None:
    """Print all rows in the workbook."""
    if not os.path.exists(path):
        print(f"No workbook found at {path}.")
        return
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        print("\t".join(str(cell) for cell in row))


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Add or list rows in an Excel workbook"
    )
    parser.add_argument(
        "--file", default=DEFAULT_FILE, help="Excel file to operate on"
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    add_parser = subparsers.add_parser("add", help="Add a row to the workbook")
    add_parser.add_argument("name", help="Name value")
    add_parser.add_argument("age", type=int, help="Age value")

    subparsers.add_parser("list", help="List rows in the workbook")

    args = parser.parse_args(argv)

    if args.command == "add":
        append_row(args.file, args.name, args.age)
    elif args.command == "list":
        print_rows(args.file)


if __name__ == "__main__":
    main()
